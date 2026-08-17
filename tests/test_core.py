from pathlib import Path

import pytest
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from pypdf import PdfReader, PdfWriter

from pdf2excel.export import export_xlsx
from pdf2excel.extraction import InvalidPasswordError, PasswordRequiredError, extract_pdf
from pdf2excel.corrections import apply_instruction
from pdf2excel.mapping import MappingStore, apply_mapping, suggest_mapping
from pdf2excel.models import ExtractionResult, ExtractedTable, MappingSpec
from pdf2excel.validation import validate
from pdf2excel.tabular import export_csv, export_text, read_excel_headers


def test_suggestions_are_reviewable_not_final():
    result = suggest_mapping(["Invoice No", "Quantity", "Customer Name"], ["Invoice Number", "Qty", "Party"])
    assert result == {"Invoice No": ["Invoice Number"], "Quantity": ["Qty"], "Customer Name": ["Party"]}


def test_combined_and_blank_mapping():
    rows = [{"First": "Ada", "Last": "Lovelace", "Amount": "1,250.00"}]
    spec = MappingSpec({"Customer": ["First", "Last"], "Tax": ["Amount"], "GSTIN": []})
    assert apply_mapping(rows, spec) == [{"Customer": "Ada Lovelace", "Tax": "1,250.00", "GSTIN": ""}]


def test_mapping_round_trip(tmp_path: Path):
    store = MappingStore(tmp_path / "mapping.json")
    spec = MappingSpec({"Invoice No": ["Invoice Number"]})
    store.save(["Invoice Number"], ["Invoice No"], spec)
    assert store.load(["Invoice Number"], ["Invoice No"]) == spec


def test_validation_preserves_duplicates_and_flags_bad_values():
    rows = [{"Invoice No": "", "Quantity": "ten", "Invoice Date": "bad"}] * 2
    extraction = ExtractionResult(list(rows[0]), rows, [ExtractedTable(1, 1, list(rows[0]), [])])
    messages = [i.message for i in validate(extraction, rows)]
    assert any("duplicate" in m for m in messages)
    assert any("non-numeric" in m for m in messages)
    assert any("blank" in m for m in messages)


def test_export_types_and_layout(tmp_path: Path):
    out = tmp_path / "out.xlsx"
    rows = [{"Invoice No": "0001", "Invoice Date": "15-08-2026", "Quantity": "2", "Total Amount": "1,250.00"}]
    export_xlsx(rows, list(rows[0]), out)
    wb = load_workbook(out)
    ws = wb.active
    assert ws["A2"].value == "0001" and ws["A2"].number_format == "@"
    assert ws["B2"].value.isoformat().startswith("2026-08-15")
    assert ws["C2"].value == 2.0 and ws["D2"].value == 1250.0
    assert ws.freeze_panes == "A2" and ws.auto_filter.ref == "A1:D2"


def test_real_pdf_table_extraction_across_pages(tmp_path: Path):
    pdf = tmp_path / "sample.pdf"
    doc = SimpleDocTemplate(str(pdf), pagesize=letter)
    data = [["Invoice Number", "Date", "Qty", "Amount"], ["INV001", "15-08-2026", "2", "1,250.00"]]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
    ]))
    doc.build([table])
    result = extract_pdf(pdf, allow_ocr=False)
    assert result.columns == data[0]
    assert result.rows == [dict(zip(data[0], data[1]))]


def test_correction_commands():
    columns = ["First", "Last", "Qty"]
    rows = [{"First": "Ada", "Last": "Lovelace", "Qty": "2"}]
    columns, rows, _ = apply_instruction(columns, rows, "rename Qty to Quantity")
    assert columns == ["First", "Last", "Quantity"]
    columns, rows, _ = apply_instruction(columns, rows, "merge First + Last into Name")
    assert columns == ["Name", "Quantity"] and rows[0]["Name"] == "Ada Lovelace"


def test_csv_text_and_template_headers(tmp_path: Path):
    columns = ["Code", "Amount"]
    rows = [{"Code": "001", "Amount": "1,250.00"}]
    csv_path, txt_path = tmp_path / "data.csv", tmp_path / "data.txt"
    export_csv(rows, columns, csv_path); export_text(rows, columns, txt_path)
    assert csv_path.read_text(encoding="utf-8-sig").splitlines()[1] == '001,"1,250.00"'
    assert txt_path.read_text(encoding="utf-8").splitlines()[0] == "Code\tAmount"
    template = tmp_path / "template.xlsx"; export_xlsx([], columns, template)
    assert read_excel_headers(template) == (columns, "Converted Data")


def test_password_protected_pdf_requests_and_accepts_password(tmp_path: Path):
    plain = tmp_path / "plain.pdf"
    encrypted = tmp_path / "protected.pdf"
    doc = SimpleDocTemplate(str(plain), pagesize=letter)
    table = Table([["Invoice", "Amount"], ["A-1", "10.00"]])
    table.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 1, colors.black)]))
    doc.build([table])
    reader = PdfReader(plain); writer = PdfWriter(); writer.append_pages_from_reader(reader); writer.encrypt("secret")
    with encrypted.open("wb") as stream:
        writer.write(stream)
    with pytest.raises(PasswordRequiredError):
        extract_pdf(encrypted, allow_ocr=False)
    with pytest.raises(InvalidPasswordError):
        extract_pdf(encrypted, allow_ocr=False, password="wrong")
    result = extract_pdf(encrypted, allow_ocr=False, password="secret")
    assert result.rows[0]["Invoice"] == "A-1"


def test_bundled_ocr_dependencies_process_scanned_pdf(tmp_path: Path):
    image = Image.new("RGB", (1800, 600), "white")
    draw = ImageDraw.Draw(image)
    font = ImageFont.truetype("arial.ttf", 72)
    draw.text((90, 100), "Invoice Amount", fill="black", font=font)
    draw.text((90, 280), "INV001 1250.00", fill="black", font=font)
    scanned = tmp_path / "scan.pdf"
    image.save(scanned, "PDF", resolution=300)
    result = extract_pdf(scanned, mode="ocr")
    assert result.used_ocr and result.rows
    assert "INV001" in " ".join(result.rows[0].values())
