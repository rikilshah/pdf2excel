from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

IDENTIFIER_HINTS = ("invoice", "gstin", "gst no", "hsn", "code", "number", " no")
NUMERIC_HINTS = ("quantity", "qty", "amount", "value", "cgst", "sgst", "igst", "tax", "total")


def _typed(column: str, value: str):
    if not value:
        return None, None
    key = column.lower()
    if "date" in key:
        for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(value, fmt).date(), "yyyy-mm-dd"
            except ValueError:
                pass
        return value, "@"
    if any(h in key for h in IDENTIFIER_HINTS):
        return value, "@"
    if any(h in key for h in NUMERIC_HINTS) and re.fullmatch(r"[-+]?\s*(?:\d{1,3}(?:,\d{3})*|\d+)(?:\.\d+)?", value):
        clean = value.replace(",", "")
        return float(clean), "#,##0.00" if "." in clean else "#,##0"
    return value, None


def export_xlsx(rows: list[dict[str, str]], columns: list[str], output: Path) -> int:
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Converted Data"
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row in rows:
        ws.append([row.get(c, "") for c in columns])
        for idx, col in enumerate(columns, 1):
            cell = ws.cell(ws.max_row, idx)
            cell.value, fmt = _typed(col, row.get(col, ""))
            if fmt:
                cell.number_format = fmt
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(1, ws.max_row)}"
    for idx, col in enumerate(columns, 1):
        sample = [str(ws.cell(r, idx).value or "") for r in range(1, min(ws.max_row, 200) + 1)]
        ws.column_dimensions[get_column_letter(idx)].width = min(50, max(10, max(map(len, sample), default=len(col)) + 2))
    wb.save(output)
    # Reopen as a structural integrity check before reporting success.
    check = load_workbook(output, read_only=True, data_only=False)
    try:
        sheet = check["Converted Data"]
        if sheet.max_row != len(rows) + 1 or [c.value for c in next(sheet.iter_rows(max_row=1))] != columns:
            raise RuntimeError("Export verification failed: workbook dimensions or headers differ.")
    finally:
        check.close()
    return len(rows)
