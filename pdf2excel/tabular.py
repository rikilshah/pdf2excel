from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

from .export import export_xlsx


def export_csv(rows: list[dict[str, str]], columns: list[str], output: Path) -> int:
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", newline="", encoding="utf-8-sig") as stream:
        writer = csv.DictWriter(stream, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    return len(rows)


def export_text(rows: list[dict[str, str]], columns: list[str], output: Path, delimiter: str = "\t") -> int:
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", newline="", encoding="utf-8") as stream:
        writer = csv.DictWriter(stream, fieldnames=columns, delimiter=delimiter, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)
    return len(rows)


def export_data(rows: list[dict[str, str]], columns: list[str], output: Path) -> int:
    suffix = output.suffix.lower()
    if suffix == ".xlsx":
        return export_xlsx(rows, columns, output)
    if suffix == ".csv":
        return export_csv(rows, columns, output)
    if suffix in {".txt", ".tsv"}:
        return export_text(rows, columns, output)
    raise ValueError("Export format must be .xlsx, .csv, .txt, or .tsv")


def read_excel_headers(path: Path, sheet_name: str | None = None, header_row: int = 1) -> tuple[list[str], str]:
    if not path.is_file():
        raise FileNotFoundError(path)
    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        selected = sheet_name or wb.sheetnames[0]
        if selected not in wb.sheetnames:
            raise ValueError(f"Worksheet not found: {selected}")
        ws = wb[selected]
        values = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True), ())
        headers = [str(v).strip() for v in values if v is not None and str(v).strip()]
        if not headers:
            raise ValueError(f"No column headers found in row {header_row} of '{selected}'.")
        if len(headers) != len(set(headers)):
            raise ValueError("The Excel header row contains duplicate names.")
        return headers, selected
    finally:
        wb.close()

